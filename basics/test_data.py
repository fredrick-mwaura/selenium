from faker import Faker
from openpyxl import Workbook ##working with xlsx

fake = Faker()

# print(fake.name())
# print(fake.email())
# print(fake.address())


wb = Workbook()
ws = wb.active

for i in range(1, 11):
    for j in range(1, 4):
        ws.cell(row=i, column=1).value = fake.name()
        ws.cell(row=i, column=2).value = fake.email()
        ws.cell(row=i, column=3).value = fake.address()
    print(fake.name())
wb.save("test.xlsx")